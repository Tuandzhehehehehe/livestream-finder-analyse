"""
database/livestream_repository.py — Database & Excel Persistence Layer
========================================================================
"""

import os
from openpyxl import load_workbook, Workbook
# pyrefly: ignore [missing-import]
from sqlalchemy import select, update, delete
# pyrefly: ignore [missing-import]
from sqlalchemy.dialects.sqlite import insert as sqlite_insert
from database.db import engine, livestreams

EXCEL_PATH = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "data", "livestreams.xlsx"))
EXCEL_HEADERS = ['Tên', 'Score', 'Priority', 'Buyer Persona', 'Industry', 'Suggested Comment', 'Location', 'Content', 'Ngày', 'YouTube', 'Meetup', 'X', 'TikTok', 'Eventbrite', 'LinkedIn']


def save_to_excel(event: dict) -> bool:
    """
    Lưu thông tin livestream vào file Excel kèm theo chỉ số đánh giá tiềm năng.
    Cột định dạng: Tên, Score, Priority, Buyer Persona, Industry, Suggested Comment, Location, Content, Ngày, YouTube, Meetup, X, TikTok, Eventbrite, LinkedIn
    """
    try:
        os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)

        headers = EXCEL_HEADERS

        if os.path.exists(EXCEL_PATH):
            try:
                wb = load_workbook(EXCEL_PATH)
                ws = wb.active
                # Cập nhật header nếu file Excel hiện tại chưa có các cột mới
                if ws.max_column < len(headers):
                    ws.delete_rows(1, ws.max_row)
                    ws.append(headers)
            except Exception as e:
                print(f"[Excel Repair] File Excel bị lỗi ({e}) - tự động khởi tạo file mới...")
                wb = Workbook()
                ws = wb.active
                ws.title = "Livestreams"
                ws.append(headers)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Livestreams"
            ws.append(headers)

        url = event.get("url", "").strip()
        if not url:
            return False

        url_exists = False
        for row in range(2, ws.max_row + 1):
            for col in range(10, 16):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val and str(cell_val).strip() == url:
                    url_exists = True
                    break
            if url_exists:
                break

        if url_exists:
            return False

        title = event.get("title", "")
        score = event.get("score", 0)
        priority = event.get("priority", "Low")
        buyer_persona = event.get("buyer_persona", "")
        industry = event.get("industry", "")
        suggested_comment = event.get("suggested_comment", "")
        location = event.get("platform", "")
        content = event.get("description", "")
        date = event.get("scheduled_start_time") or event.get("start_time") or ""

        row_data = [title, score, priority, buyer_persona, industry, suggested_comment, location, content, date, "", "", "", "", "", ""]

        platform = str(event.get("platform", "")).lower().strip()
        if "youtube" in platform:
            row_data[9] = url
        elif "meetup" in platform:
            row_data[10] = url
        elif "x" in platform or "twitter" in platform:
            row_data[11] = url
        elif "tiktok" in platform:
            row_data[12] = url
        elif "eventbrite" in platform:
            row_data[13] = url
        elif "linkedin" in platform:
            row_data[14] = url

        ws.append(row_data)
        wb.save(EXCEL_PATH)
        return True
    except Exception as e:
        print(f"❌ Excel save error: {e}")
        return False
def save_event(event: dict) -> bool:
    try:
        with engine.begin() as conn:
            stmt = sqlite_insert(livestreams).values(
                title=event["title"],
                platform=event.get("platform"),
                description=event.get("description"),
                url=event["url"],
                keyword=event.get("keyword"),
                status=event.get("status"),
                start_time=event.get("start_time"),
                scheduled_start_time=event.get("scheduled_start_time"),
                actual_start_time=event.get("actual_start_time"),
                actual_end_time=event.get("actual_end_time"),
                score=event.get("score"),
                industry=event.get("industry"),
                language=event.get("language"),
                buyer_persona=event.get("buyer_persona"),
                priority=event.get("priority"),
                interaction_tip=event.get("interaction_tip"),
                suggested_comment=event.get("suggested_comment"),
            ).on_conflict_do_nothing(index_elements=["url"])

            res = conn.execute(stmt)
            if res.rowcount > 0:
                save_to_excel(event)
                return True
            return False
    except Exception as e:
        print(f"❌ Save event error: {e}")
        return False


def get_all_events():
    with engine.connect() as conn:
        return conn.execute(select(livestreams)).fetchall()


def get_event_by_id(event_id: int):
    with engine.connect() as conn:
        return conn.execute(select(livestreams).where(livestreams.c.id == event_id)).fetchone()


def get_event_by_url(url: str):
    with engine.connect() as conn:
        return conn.execute(select(livestreams).where(livestreams.c.url == url)).fetchone()


def update_classification_by_url(url: str, industry: str, language: str, buyer_persona: str, score: int):
    with engine.begin() as conn:
        conn.execute(update(livestreams).where(livestreams.c.url == url).values(
            industry=industry, language=language, buyer_persona=buyer_persona, score=score or 0
        ))


def update_suggested_comment_by_url(url: str, suggested_comment: str):
    with engine.begin() as conn:
        conn.execute(update(livestreams).where(livestreams.c.url == url).values(suggested_comment=suggested_comment))


def delete_event_by_url(url: str) -> bool:
    with engine.begin() as conn:
        res = conn.execute(delete(livestreams).where(livestreams.c.url == url))
        return res.rowcount > 0